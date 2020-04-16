import yfinance as yf
import datetime
import calendar
import pandas as pd
from openpyxl import load_workbook

def getTickersData(startDate):
    tickers = 'BRL=X AUD=X MXN=X'
    data = yf.download(tickers=tickers, start=startDate, group_by="ticker", interval="1m")
    return data.to_dict('series')

def writeExcel(filename, dictInsData, sheet, date):
    wb = load_workbook(filename, False)
    ws = wb.get_sheet_by_name(sheet)

    cellPos = ws.max_row

    for i, j in dictInsData:
        if j =='Close':
            for index in dictInsData[i,j].index:
                if (dictInsData[i,j][index] > 0) and (str(index).find(' 18:00') > 0):
                    cellPos = cellPos + 1
                    print(i)
                    print(j)
                    print(index)
                    print(dictInsData[i,j][index])
                    ws['A' + str(cellPos)].value = i
                    ws['B' + str(cellPos)].value = index
                    ws['C' + str(cellPos)].value = dictInsData[i,j][index]
                    ws['D' + str(cellPos)].value = date

    wb.save(filename)

    wb.close()

def get_date():
    todayDate = datetime.datetime.today()

    if calendar.weekday(todayDate.year, todayDate.month, todayDate.day) == 6:
        previousDate = todayDate - datetime.timedelta(days=2)
    elif calendar.weekday(todayDate.year, todayDate.month, todayDate.day) == 0:
        previousDate = todayDate - datetime.timedelta(days=3)
    else:
        previousDate = todayDate - datetime.timedelta(days=1)

    return (str(previousDate.year) + '-' + str(previousDate.month) + '-' + str(previousDate.day))

if __name__ == '__main__':
    #Carregando Variaveis
    excelFilename = 'RawDataYahoo.xlsx'
    yesterdayDate = get_date()
    dictData = getTickersData(yesterdayDate)
    dateToday = datetime.datetime.now()
    writeExcel(filename=excelFilename, dictInsData=dictData, sheet='YahooFX', date=dateToday)