import time
import pandas as pd
import threading
import _thread
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from bs4 import BeautifulSoup
from openpyxl import load_workbook
import datetime

def executeRobotBroadcast(url):
    driver.get(url)
    time.sleep(2)
    element = driver.find_element_by_xpath("//table[@class='table table-sm table-hover']")
    htmlContent = element.get_attribute('outerHTML')

    soup = BeautifulSoup(htmlContent, 'html.parser')
    table = soup.find(name='table')

    df = pd.read_html(str(table))[0].head(18)
    df.reset_index(drop=True, inplace=True)
    
    dictIndicadores = df.to_dict('records')

    return dictIndicadores

def executeRobotAmbito(url):
    driver.get(url)
    time.sleep(2)
    element = driver.find_element_by_xpath("//div[@class='indicador-historico-general']//table[@class='table']")
    htmlContent = element.get_attribute('outerHTML')

    table = BeautifulSoup(htmlContent, 'html.parser')
    table = str(table).replace(",", ".")
    
    df = pd.read_html(str(table))
   
    return df

def writeExcel(filename, dictInsData, sheet, date):
    wb = load_workbook(filename, False)
    ws = wb.get_sheet_by_name(sheet)

    cellPos = ws.max_row
    
    if sheet == "Broadcast":
        for data in dictInsData:
            if data[0] == 'Risco País - Brasil (pontos)':
                if ws['B' + str(cellPos)].value == data[1]:
                    ws['C' + str(cellPos)].value = date
                else:     
                    cellPos = cellPos + 1
                    ws['A' + str(cellPos)].value = data[0]
                    ws['B' + str(cellPos)].value = data[1]
                    ws['C' + str(cellPos)].value = date
    else:
        ws.delete_rows(2, cellPos)
        cellPos = 1
        
        for data in range(0, len(dictInsData[0]['Puntos'])):
            if dictInsData[0]['Puntos'][data] > 0 :
                cellPos = cellPos + 1
                ws['A' + str(cellPos)].value = 'Risco País - Brasil (pontos)'
                ws['B' + str(cellPos)].value = dictInsData[0]['Puntos'][data]
                ws['C' + str(cellPos)].value = dictInsData[0]['Fecha'][data]
                ws['D' + str(cellPos)].value = date

    wb.save(filename)

    wb.close()

if __name__ == '__main__':
    #Carregando Variaveis    
    urlBroadcast = 'http://cliente.estadaoconteudodados.com.br/agencia/sitebroadcast/tabelas/indicadores.html'
    urlAmbito = 'https://www.ambito.com/contenidos/riesgo-pais-brasil-historico.html'
    excelFilename = 'RawData.xlsx'

    option = Options()
    option.headless = True
    option.add_argument("--log-level=3")
    driver = webdriver.Chrome(options=option)
    
    inputQuit = 'N'

    dateToday = datetime.datetime.now()
    #print('Sincronizando horário!')
    #while dateToday.minute not in (0,15,30,45):
    #    time.sleep(25)   
    #    dateToday = datetime.datetime.now()

    #print('Horário Sincronizado ' + str(dateToday))

    try:

        while inputQuit != 'Y':
            dateToday = datetime.datetime.now()
            print('Processando Broadcast: ' + urlBroadcast)
            dictIndicadores = executeRobotBroadcast(url=urlBroadcast)
            writeExcel(filename=excelFilename, dictInsData=dictIndicadores, sheet='Broadcast', date=dateToday)
            print('Processando Ambito: ' + urlBroadcast)
            dictIndicadores = executeRobotAmbito(url=urlAmbito)
            writeExcel(filename=excelFilename, dictInsData=dictIndicadores, sheet='Ambito', date=dateToday)
            print('Sleeping...')
            time.sleep(15)
            #print('Verificando horário limite execução!')
            #if dateToday.hour >= 20:
            #    inputQuit = 'Y'
            #    driver.quit()
    except KeyboardInterrupt:
        driver.quit()