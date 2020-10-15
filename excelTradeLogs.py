from openpyxl import load_workbook
import datetime


def writeExcel(dictInsData):
    #print('Writing excel.')
    excelFilename = 'quantTrades.xlsx'
    sheet = 'Trades'
    dateToday = datetime.datetime.now()

    wb = load_workbook(excelFilename, False)
    ws = wb.get_sheet_by_name(sheet)

    cellPos = ws.max_row
    
    cellPos = cellPos + 1
    ws['A' + str(cellPos)].value = dictInsData['CorrelationId']
    ws['B' + str(cellPos)].value = dictInsData['Ticker']
    ws['C' + str(cellPos)].value = dictInsData['Units']
    ws['D' + str(cellPos)].value = dictInsData['UnitValue']
    ws['E' + str(cellPos)].value = dictInsData['OrderType']
    ws['F' + str(cellPos)].value = dateToday
    ws['G' + str(cellPos)].value = dictInsData['Control']
        
    wb.save(excelFilename)

    wb.close()