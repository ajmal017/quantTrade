import time
import pandas as pd
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from bs4 import BeautifulSoup
from openpyxl import load_workbook
import datetime

def executeRobotCME(url):
    driver.get(url)
    time.sleep(2)
    element = driver.find_element_by_xpath("//table[@id='quotesFuturesProductTable1']")
    htmlContent = element.get_attribute('outerHTML')

    soup = BeautifulSoup(htmlContent, 'html.parser')
    table = soup.find(name='table')

    df = pd.read_html(str(table))[0].head(1)
    df.reset_index(drop=True, inplace=True)
    
    dictIndicadores = df.to_dict('records')

    return dictIndicadores

def writeExcel(filename, dictInsData, sheet, date):
    wb = load_workbook(filename, False)
    ws = wb.get_sheet_by_name(sheet)

    cellPos = ws.max_row
    
    for data in dictInsData:
        if ws['B' + str(cellPos)].value == data['Last']:
            ws['C' + str(cellPos)].value = date
        else:     
            cellPos = cellPos + 1
            ws['A' + str(cellPos)].value = data['Month']
            ws['B' + str(cellPos)].value = data['Last']
            ws['C' + str(cellPos)].value = date
        
    wb.save(filename)

    wb.close()

if __name__ == '__main__':
    #Carregando Variaveis
    urlCME = 'https://www.cmegroup.com/trading/fx/emerging-market/brazilian-real.html'
    excelFilename = 'RawDataCME.xlsx'

    option = Options()
    option.headless = True
    driver = webdriver.Chrome(options=option)
    
    try:
        while True:
            dateToday = datetime.datetime.now()
            print('Processando CME: ' + urlCME)
            dictIndicadores = executeRobotCME(url=urlCME)
            writeExcel(filename=excelFilename, dictInsData=dictIndicadores, sheet='CME', date=dateToday)
    except KeyboardInterrupt:
        driver.quit()