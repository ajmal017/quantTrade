import time
import pyperclip
import pandas as pd
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.keys import Keys
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from decimal import Decimal
import datetime

def executeRobotPtax(url, baseDate):
    driver.get(url)
    time.sleep(4)
    #print(driver.page_source)
    driver.find_element_by_xpath("//input[@id='RadOpcao' and @value='3']").click()
    #pyperclip.copy(baseDate)
    driver.find_elements_by_id('DATAINI')[0].clear()
    #driver.find_elements_by_id('DATAINI')[0].send_keys(Keys.CONTROL,'v')
    driver.find_elements_by_id('DATAINI')[0].send_keys(baseDate.replace('/',''))
    driver.find_elements_by_xpath("//input[@title='Pesquisar']")[0].click()
    
    msgErro = driver.find_elements_by_xpath("//*[@class='msgErro']")
    if len(msgErro) != 0: 
        return -1

    time.sleep(4)
    element = driver.find_element_by_xpath("//table[@class='tabela']")
    htmlContent = element.get_attribute('outerHTML')

    soup = BeautifulSoup(htmlContent, 'html.parser')
    table = soup.find(name='table')
    
    df = pd.read_html(str(table))[0]
    df.columns = df.columns.get_level_values(2)
    df.columns = ['Hora', 'Tipo', 'Compra', 'Venda', 'Trash', 'Trash', 'Trash']
    df.drop(['Trash'], axis=1, inplace=True)
    df.drop(df.tail(1).index,inplace=True)
    
    return df

def writeExcel(filename, dfData, sheet, date):
    wb = load_workbook(filename, False)
    ws = wb.get_sheet_by_name(sheet)

    cellPos = ws.max_row
    
    if cellPos > 1:
        lastCell = 'E' + str(cellPos)
        cellsSort = 'E2:' + lastCell

        ws.auto_filter.add_sort_condition(cellsSort, descending=True)
    
        startRow = ws.max_row

        if startRow > 20:
            lastRow = ws.max_row - 15
        else:
            lastRow = 1

        for i in range(startRow, lastRow, -1):
            cellDate = 'E' + str(i)
            cellContent = ws[cellDate].value
        
            if(type(cellContent) is str):
                testDate1 = datetime.datetime.strptime(cellContent, '%d/%m/%Y')
        
            testDate2 = datetime.datetime.strptime(date, '%d/%m/%Y')

            if testDate1.date() == testDate2.date():
                ws.delete_rows(i, 1)

    cellPos = ws.max_row

    for i in range(len(dfData)):
        cellPos = cellPos + 1
        ws['A' + str(cellPos)].value = dfData['Hora'][i]
        ws['B' + str(cellPos)].value = dfData['Tipo'][i]
        ws['C' + str(cellPos)].value = Decimal(dfData['Compra'][i])/10000
        ws['D' + str(cellPos)].value = Decimal(dfData['Venda'][i])/10000
        ws['E' + str(cellPos)].value = date
        
    wb.save(filename)

    wb.close()

def readExcel(filename, sheet):
    wb = load_workbook(filename, False)
    ws = wb.get_sheet_by_name(sheet)

    cellName = 'A2'
    baseDate = ws[cellName].value
    wb.close()
    
    return baseDate

def updateExcel(filename, sheet, date):
    wb = load_workbook(filename, False)
    ws = wb.get_sheet_by_name(sheet)

    cellName = 'A2'
    ws[cellName].value = date

    wb.save(filename)
    wb.close()

if __name__ == '__main__':
    #Carregando Variaveis
    urlPtax = 'https://ptax.bcb.gov.br/ptax_internet/consultaBoletim.do?method=exibeFormularioConsultaBoletim'
    excelFilename = 'RawDataPtax.xlsx'

    while True:
        try:
            dateToday = datetime.datetime.now()
            option = Options()
            option.headless = True
            option.add_argument("--log-level=3")
            driver = webdriver.Chrome(options=option)
    
            refDate =  readExcel(filename=excelFilename, sheet='DataBase')
            if(type(refDate) is str):
                refDate = datetime.datetime.strptime(refDate, '%d/%m/%Y')
            
            refDateFmt = str(refDate.day).zfill(2) + '/' + str(refDate.month).zfill(2) + '/' + str(refDate.year)
            print('Executando Data Base: ' + refDateFmt)

            print('Obtendo Ptax em: ' + urlPtax)
            dfPtax = executeRobotPtax(url=urlPtax, baseDate=refDateFmt)
            
            if isinstance(dfPtax, pd.DataFrame):
                writeExcel(filename=excelFilename, dfData=dfPtax, sheet='Ptax', date=refDateFmt)
            
                if any('Fech' in s for s in dfPtax['Tipo']):
                    print('Encerrando processamento dia: ' + refDateFmt)
                    refDate = refDate + datetime.timedelta(days=1)
                    refDateFmt = str(refDate.day).zfill(2) + '/' + str(refDate.month).zfill(2) + '/' + str(refDate.year)
                    print('Atualiando data base para: ' + refDateFmt)
                    updateExcel(filename=excelFilename, sheet='DataBase', date=refDateFmt)
            else:
                print('Data Base não possui dados de Ptax!')
                if dateToday.date() > refDate.date():
                    refDate = refDate + datetime.timedelta(days=1)
                    refDateFmt = str(refDate.day).zfill(2) + '/' + str(refDate.month).zfill(2) + '/' + str(refDate.year)
                    print('Atualiando data base para: ' + refDateFmt)
                    updateExcel(filename=excelFilename, sheet='DataBase', date=refDateFmt)

            driver.quit()
            print('Sleeping...')
            time.sleep(60)
        except:
            print("Ocorreu um problema, reiniciando!")
            driver.quit()